#importações
from selenium import webdriver
from selenium.webdriver.common.by import By
import requests 
import pandas as pd
import os
from time import *
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import xlrd
import numpy as np
import mysql.connector
from mysql.connector import connection
from unidecode import unidecode
import ssl
from database import MariaDB
from webdriver_manager.microsoft import EdgeChromiumDriverManager


#passando os parametros para apagar arquivos
def apagar_arquivo():
    path = r"C:\Users\Plugify\Documents\Repositorios\Crawler_CMED\\"
    dir = os.listdir(path)
    for file in dir:
        if file == "arquivo.xls":
            try:
                os.remove(path+file)
                print('arquivo ("arquivo.xls") removido...')
            except Exception as e:
                print(f'Erro ao remover o arquivo "{file}": {str(e)}')  
                
        elif file == "arquivo_convertido.xlsx":
            try:
                os.remove(path+file)
                print('arquivo ("arquivo_convertido.xlsx") removido...')
            except:
                print('erro ao remover o arquivo ("arquivo_convertido.xlsx")...') 
                
        elif file == "arquivo_convertido_tratado.xlsx":
            try:
                os.remove(path+file)
                print('arquivo ("arquivo_convertido_tratado.xlsx") removido...')
            except:
                print('erro ao remover o arquivo ("arquivo_convertido_tratado.xlsx")...')  
                


apagar_arquivo()


#realizando o scrapping
#pra achar o xpath inspecionar a página e ir no div qu quer e com o botao direito->copy->xpath
def scraping_inclusao():

    #try:

    print('url e credenciais...')
    url = 'https://www.gov.br/anvisa/pt-br/assuntos/medicamentos/cmed/precos'

    print('instanciando webdriver...\n')
    navegador = webdriver.Edge()

    print('acessando url via get...\n')
    navegador.get(url)

    print('minimizando janela...\n')
    navegador.minimize_window()
        
    print('Clicando para dowloand...\n')

    elemento = navegador.find_element(By.XPATH,'//*[@id="db63c5be-7407-4436-aa28-4c16db6b6162"]/div/div/div/div[2]/a')
    print('elemento...\n', elemento)
    url_download = elemento.get_attribute("href")
    print('url_download..\n', url_download)
    r = requests.get(url_download)
    print('r..\n',r)

    with open("arquivo.xls", "wb") as arquivo:
        arquivo.write(r.content)

    print('aguardando 10 segundos para completar download do arquivo...\n')
    sleep(10)

    print('fechando navegador...\n')
    navegador.close()
 

    #print('retorno do webscraping: \n', df)
    print('fim do scraping...\n')


#executando o download
def executa_scraping_inclusao():
    df_inclusao =  pd.DataFrame(scraping_inclusao()) #chama a execução do download
    print('rotina de scraping -> def scraping_inclusao(cmed) finalizado... ')


#executa o def
executa_scraping_inclusao()


#transformando em xlsx pra conseguir usar a biblioteca do openpxl
df = pd.read_excel(r"C:\Users\Plugify\Documents\Repositorios\crawler_CMED\arquivo.xls")
df.to_excel("arquivo_convertido.xlsx", index=False)


#colocando o arquvi pra ser lido pela biblioteca workbook para tratar e remover as linhas acima
wb = load_workbook(r"C:\Users\Plugify\Documents\Repositorios\crawler_CMED\arquivo_convertido.xlsx")
ws = wb.active


#removendo as linhas acima de substancia na coluna 1

value_to_match = "SUBSTÂNCIA"

for row in reversed(range(1, ws.max_row + 1)):
    if ws.cell(row=row, column=1).value == value_to_match:
        ws.delete_rows(1, row - 1)
        break

# Salvando a anteração em um novo arquivo de excel

new_workbook = Workbook()
new_worksheet = new_workbook.active
for row in ws:
    for cell in row:
        column_letter = get_column_letter(cell.column)
        new_worksheet[f"{column_letter}{cell.row}"] = cell.value
new_workbook.save(filename="arquivo_convertido_tratado.xlsx")


#lê o arquivo baixado e tratado joga em um df


colunas_necessarias = ['SUBSTÂNCIA','LABORATÓRIO','REGISTRO','EAN 1','PRODUTO','APRESENTAÇÃO',
                           'CLASSE TERAPÊUTICA','TIPO DE PRODUTO (STATUS DO PRODUTO)','PF 0%','PMC 0%',
                           'RESTRIÇÃO HOSPITALAR','TARJA']

df = pd.read_excel(r"C:\Users\Plugify\Documents\Repositorios\crawler_CMED\arquivo_convertido_tratado.xlsx", usecols=colunas_necessarias)
